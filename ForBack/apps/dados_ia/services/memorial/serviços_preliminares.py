import re
import io
import json
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.cell.cell import MergedCell

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s — %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("servicos_preliminares")

MAX_ITENS = 80
CAMINHO_TEMPLATE = Path(__file__).resolve().parent / "templates_excel" / "Memorial de Cálculo - Modelo.xlsx"

def _clean_text(raw):
    if not raw: return ""
    t = re.sub(r'\{[^{}]*\}', '', raw)
    t = re.sub(r'\\[a-zA-Z0-9]+\d*\.?.?x?;', ' ', t)
    t = t.replace('\\P', ' ').replace('\\p', ' ')
    t = re.sub(r'\s+', ' ', t)
    return t.strip()

def get_coluna_remocao(nome):
    nome = nome.upper()
    if "CONDULETE" in nome: return 5
    if "TOMADA" in nome: return 6
    if "INTERRUPTOR" in nome: return 7
    if "LUMIN" in nome: return 8
    if "DUTO" in nome or "FIAÇÃO" in nome or "FIACAO" in nome or "CABO" in nome: return 10
    if "CAPTAÇÃO" in nome or "CAPTACAO" in nome: return 11
    if "ATERRA" in nome: return 12
    if "QUADRO" in nome: return 13
    if "POSTE" in nome or re.search(r'\bP\d+\b', nome): return 14
    if "CAVALETE" in nome: return 15
    if "RESERVAT" in nome: return 16
    if "REGISTRO" in nome: return 17
    if "VÁLVULA" in nome or "VALVULA" in nome: return 18
    if "TORNEIRA" in nome: return 19
    if "CALHA" in nome: return 21
    if "CAIXA" in nome: return 22
    if "DRENO" in nome: return 23
    if "PORTA" in nome: return 24
    if "JANELA" in nome or "ESQUADRIA" in nome: return 25
    if "TELHA" in nome: return 26
    if "TRAMA" in nome: return 27
    if "TESOURA" in nome: return 28
    return "GENERICO" 

def get_coluna_demolicao(nome):
    nome = nome.upper()
    if "PISO" in nome: return 5
    if "RODAPÉ" in nome or "RODAPE" in nome: return 6
    if "AZULEJO" in nome: return 7
    if "FORRO" in nome: return 8
    if "ALVENARIA" in nome or "PLATIBANDA" in nome: return 10
    if "FUNDAÇÃO" in nome or "FUNDACAO" in nome: return 11
    if "PILAR" in nome or "POSTE" in nome or re.search(r'\bP\d+\b', nome): return 12
    if "VIGA" in nome: return 13
    if "LAJE" in nome: return 14
    return None

def _extrair_servicos_agrupados(textos):
    demolicoes_final = []
    remocoes_final = []
    
    seen_dem = set()
    seen_rem = set()
    
    counter_dem = {}
    counter_rem = {}
    
    ruidos = ['PLANTA', 'APENAS', 'LEGENDA', 'ESCALA']

    for t in textos:
        c = t.get('conteudo', '')
        if not c: continue
        
        c_upper = c.upper()
        if any(r in c_upper for r in ruidos): continue

        raw_pos = t.get('posicao', [0, 0])
        pos_xy = tuple(round(p, 1) for p in raw_pos[:2])

        conteudo_limpo = _clean_text(c)
        
        match_metragem = re.search(r'(\d+[.,]?\d*)\s*(m|metros)\b', conteudo_limpo, re.IGNORECASE)
        metragem_val = float(match_metragem.group(1).replace(',','.')) if match_metragem else None

        is_demolicao = any(k in c_upper for k in ['DEMOL', 'DEMOLIÇÃO', 'DEMOLIDA', 'DEMOLIDO'])
        is_remocao = any(k in c_upper for k in ['REMOV', 'REMOÇÃO', 'RETIRAR', 'REMOVIDO', 'REMOVIDA'])

        if is_demolicao:
            if c_upper.strip() not in ['A DEMOLIR', 'DEMOLIÇÃO', 'DEMOLIR', 'A SER DEMOLIDA', 'A SER DEMOLIDO']:
                sujeito = re.sub(r'\bA\s+(SER\s+)?DEMOL\w+\b(\s*/\s*REATERRAR)?', '', conteudo_limpo, flags=re.IGNORECASE)
                sujeito = re.sub(r'\(\s*\)', '', sujeito)
                sujeito = re.sub(r'-\s*$', '', sujeito).strip()
                
                if len(sujeito) > 1:
                    nome_padronizado = f"{sujeito.upper()} A DEMOLIR"
                    chave = (nome_padronizado, pos_xy)
                    
                    if chave not in seen_dem:
                        seen_dem.add(chave)
                        counter_dem[nome_padronizado] = counter_dem.get(nome_padronizado, 0) + 1
                        idx = counter_dem[nome_padronizado]
                        
                        nome_exibicao = nome_padronizado if idx == 1 else f"{nome_padronizado} ({idx})"
                        
                        is_continuo = any(k in nome_padronizado for k in ['DUTO', 'FIAÇÃO', 'CABO', 'ALVENARIA', 'PLATIBANDA', 'PISO'])
                        is_pilar_estrutural = any(k in nome_padronizado for k in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'PILAR', 'POSTE', 'FUNDAÇÃO', 'VIGA', 'LAJE'])
                        
                        if is_continuo:
                            valor = metragem_val if metragem_val is not None else ""
                        elif is_pilar_estrutural:
                            valor = ""
                        else:
                            valor = 1
                            
                        demolicoes_final.append((nome_exibicao, valor))

        if is_remocao and not is_demolicao:
            if c_upper.strip() not in ['A REMOVER', 'REMOÇÃO', 'REMOVER', 'A RETIRAR', 'RETIRAR']:
                sujeito = re.sub(r'\b(A\s+(SER\s+)?(REMOV\w+|RETIRAR\w*)|REMOÇÃO\s+DE|RETIRADA\s+DE)\b', '', conteudo_limpo, flags=re.IGNORECASE)
                sujeito = re.sub(r'\(\s*\)', '', sujeito)
                sujeito = re.sub(r'-\s*$', '', sujeito).strip()
                
                if len(sujeito) > 1:
                    nome_padronizado = f"{sujeito.upper()} A REMOVER"
                    chave = (nome_padronizado, pos_xy)
                    
                    if chave not in seen_rem:
                        seen_rem.add(chave)
                        counter_rem[nome_padronizado] = counter_rem.get(nome_padronizado, 0) + 1
                        idx = counter_rem[nome_padronizado]
                        
                        nome_exibicao = nome_padronizado if idx == 1 else f"{nome_padronizado} ({idx})"
                        
                        is_continuo = any(k in nome_padronizado for k in ['DUTO', 'FIAÇÃO', 'FIACAO', 'CABO'])
                        
                        if is_continuo:
                            valor = metragem_val if metragem_val is not None else ""
                        else:
                            valor = 1
                            
                        remocoes_final.append((nome_exibicao, valor))

    return remocoes_final, demolicoes_final

def escrever(ws, linha, coluna, valor):
    if valor is None or valor == "": return 
    
    celula = ws.cell(row=linha, column=coluna)
    if isinstance(celula, MergedCell):
        for rng in ws.merged_cells.ranges:
            if celula.coordinate in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = valor
                return
    else:
        celula.value = valor

def extrair_servicos_preliminares_para_xlsx(dados_json: dict, template_path: str = None) -> bytes:
    log.info("=" * 60)
    log.info("INÍCIO")
    log.info("=" * 60)

    if 'entidades' not in dados_json:
        for val in dados_json.values():
            if isinstance(val, dict) and 'entidades' in val:
                dados_json = val; break
            if isinstance(val, str):
                try:
                    parsed = json.loads(val)
                    if 'entidades' in parsed: dados_json = parsed; break
                except: pass

    textos = dados_json.get('textos', []) or [e for e in dados_json.get('entidades', []) if e.get('tipo') in ('MTEXT', 'TEXT')]
    
    itens_remocao, itens_demolicao = _extrair_servicos_agrupados(textos)

    tpl = Path(template_path) if template_path else CAMINHO_TEMPLATE
    wb = openpyxl.load_workbook(str(tpl))
    ws = wb['Serviços Preliminares']
    
    linha_inicio_remocoes = None
    linha_inicio_demolicoes = None
    
    for r in range(1, max(ws.max_row + 1, 200)):
        val_a = str(ws.cell(row=r, column=1).value or "").strip().lower()
        val_b = str(ws.cell(row=r, column=2).value or "").strip().lower()
        texto_linha = val_a + " " + val_b
        
        if "remoções" in texto_linha and "1.7" in texto_linha:
            linha_inicio_remocoes = r + 5
        elif "demolições" in texto_linha and "1.8" in texto_linha:
            linha_inicio_demolicoes = r + 6

    if linha_inicio_remocoes:
        for i, (nome, valor) in enumerate(itens_remocao[:MAX_ITENS]):
            linha = linha_inicio_remocoes + i
            ws.cell(row=linha, column=2).value = nome 
            
            col_id = get_coluna_remocao(nome)
            if col_id == "GENERICO":
                escrever(ws, linha, 29, nome) 
                escrever(ws, linha, 31, valor)  
            else:
                escrever(ws, linha, col_id, valor)

    if linha_inicio_demolicoes:
        for i, (nome, valor) in enumerate(itens_demolicao[:MAX_ITENS]):
            linha = linha_inicio_demolicoes + i
            ws.cell(row=linha, column=2).value = nome 
            
            col_id = get_coluna_demolicao(nome)
            if col_id is not None:
                escrever(ws, linha, col_id, valor)

    output = io.BytesIO()
    wb.save(output)
    
    return output.getvalue()