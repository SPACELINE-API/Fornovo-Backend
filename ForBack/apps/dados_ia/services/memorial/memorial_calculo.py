import re
import math
import io
import json
import logging
from collections import Counter
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.cell.cell import MergedCell

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s — %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("processamento_planilhas")

MAX_AMBIENTES = 80
MAX_ITENS = 80
ESPESSURA_PADRAO = 0.15
PD_PADRAO = 3.05
RAIO_VAOS = 8.0

LINHA_INI_AMBIENTES  = 8
LINHA_INI_ESTRUTURAS = 114
LINHA_INI_ELETRICA   = 227
LINHA_INI_HIDRAULICA = 305
LINHA_INI_REDE       = 380
LINHA_INI_INCENDIO   = 458
LINHA_INI_COBERTURA  = 531
LINHA_LOG            = 588

COL_AMB_NOME  = 2
COL_AMB_C     = 5
COL_AMB_L     = 6
COL_AMB_H     = 7
COL_AMB_E     = 8
COL_AMB_AREA  = 9
COL_AMB_VAOS  = 10

COL_EST_PILAR_C = 5
COL_EST_PILAR_L = 6
COL_EST_PILAR_H = 7

COL_ELE_QUADROS  = 5
COL_ELE_CONDULE  = 6
COL_ELE_TOMADAS  = 7
COL_ELE_INTERR   = 8
COL_ELE_LUM      = 9
COL_ELE_DUTOS    = 10
COL_ELE_CABOS    = 11
COL_ELE_TIPO_AC  = 12   
COL_ELE_AC_QTD   = 14  
COL_ELE_EQUIP    = 15   
COL_ELE_EQUIP_QTD= 17   

COL_HID_CAVALETES  = 5
COL_HID_RESERVAT   = 6   
COL_HID_RESERVAT_L = 7   
COL_HID_REGISTROS  = 8
COL_HID_VALVULAS   = 9
COL_HID_TORNEIRAS  = 10
COL_HID_AF_DUTOS   = 11
COL_HID_AP_CALHAS  = 12  
COL_HID_AP_CALHAS_M= 13  
COL_HID_AP_DUTOS   = 14
COL_HID_AP_CAIXAS  = 15
COL_HID_ESG_DRENOS = 16
COL_HID_ESG_DUTOS  = 17
COL_HID_ESG_CAIXAS = 18

COL_REDE_QUADROS    = 5
COL_REDE_CONDULE    = 6
COL_REDE_TOMADAS    = 7
COL_REDE_DUTOS_TIPO = 8
COL_REDE_DUTOS_M    = 9
COL_REDE_CABOS      = 10

COL_SPDA_CAPTACAO   = 11
COL_SPDA_CONDULE    = 12
COL_SPDA_ATERR      = 13
COL_SPDA_DUTOS_TIPO = 14
COL_SPDA_DUTOS_M    = 15
COL_SPDA_CABOS      = 16

COL_INC_RES_TIPO   = 5
COL_INC_RES_QTD    = 6
COL_INC_REGISTROS  = 7
COL_INC_VALVULAS   = 8
COL_INC_DUTOS_TIPO = 9
COL_INC_DUTOS      = 10
COL_INC_HIDRANTES  = 11

COL_COB_C          = 5
COL_COB_L          = 6
COL_COB_H          = 7
COL_COB_EST_TIPO   = 8
COL_COB_EST_L      = 9
COL_COB_EST_C      = 10
COL_COB_EST_E      = 11
COL_COB_EST_AREA   = 12
COL_COB_TEL_TIPO   = 13
COL_COB_TEL_L      = 14
COL_COB_TEL_C      = 15
COL_COB_TEL_E      = 16

SOLO_LINHA_TERRAPLANAGEM = 17   
SOLO_LINHA_ATERRO        = 43  
SOLO_LINHA_ENROCAMENTO   = 69 
SOLO_LINHA_CONTENCAO     = 95  
SOLO_LINHA_TALUDAMENTO   = 121  
SOLO_LINHA_NIVELAMENTO   = 148

TERRA_COL_AMB    = 2
TERRA_COL_TIPO   = 5
TERRA_COL_H      = 9
TERRA_COL_LASTRO = 10
TERRA_COL_V      = 12

SOLO_COL_AMB = 2
SOLO_COL_I   = 5
SOLO_COL_V   = 10

CAMINHO_TEMPLATE = Path(__file__).resolve().parent / "templates_excel" / "Memorial de Cálculo - Modelo.xlsx"

def _clean_text_ambientes(raw):
    if not raw: return ""
    t = re.sub(r'\{[^{}]*\}', '', raw)
    t = re.sub(r'\\[a-zA-Z0-9]+\d*\.?.?x?;', ' ', t)
    t = re.sub(r'\s+', ' ', t)
    return t.strip().strip('\\{}()[]').strip()

def _clean_text_servicos(raw):
    if not raw: return ""
    t = re.sub(r'\{[^{}]*\}', '', raw)
    t = re.sub(r'\\[a-zA-Z0-9]+\d*\.?.?x?;', ' ', t)
    t = t.replace('\\P', ' ').replace('\\p', ' ')
    t = re.sub(r'\s+', ' ', t)
    return t.strip()

def get_coluna_remocao(nome):
    nome = nome.upper()
    if "CONDULETE" in nome: return 5
    if "TOMADA" in nome: return 6
    if "INTERRUPTOR" in nome: return 7
    if "LUMIN" in nome: return 8
    if "DUTO" in nome or "FIAÇÃO" in nome or "FIACAO" in nome or "CABO" in nome: return 10
    if "CAPTAÇÃO" in nome or "CAPTACAO" in nome: return 11
    if "ATERRA" in nome: return 12
    if "QUADRO" in nome: return 13
    if "POSTE" in nome or re.search(r'\bP\d+\b', nome): return 14
    if "CAVALETE" in nome: return 15
    if "RESERVAT" in nome: return 16
    if "REGISTRO" in nome: return 17
    if "VÁLVULA" in nome or "VALVULA" in nome: return 18
    if "TORNEIRA" in nome: return 19
    if "CALHA" in nome: return 21
    if "CAIXA" in nome: return 22
    if "DRENO" in nome: return 23
    if "PORTA" in nome: return 24
    if "JANELA" in nome or "ESQUADRIA" in nome: return 25
    if "TELHA" in nome: return 26
    if "TRAMA" in nome: return 27
    if "TESOURA" in nome: return 28
    return "GENERICO" 

def get_coluna_demolicao(nome):
    nome = nome.upper()
    if "PISO" in nome: return 5
    if "RODAPÉ" in nome or "RODAPE" in nome: return 6
    if "AZULEJO" in nome: return 7
    if "FORRO" in nome: return 8
    if "ALVENARIA" in nome or "PLATIBANDA" in nome: return 10
    if "FUNDAÇÃO" in nome or "FUNDACAO" in nome: return 11
    if "PILAR" in nome or "POSTE" in nome or re.search(r'\bP\d+\b', nome): return 12
    return None

def _extrair_servicos_agrupados(textos):
    demolicoes_final = []
    remocoes_final = []
    seen_dem = set()
    seen_rem = set()
    counter_dem = {}
    counter_rem = {}
    ruidos = ['PLANTA', 'APENAS', 'LEGENDA', 'ESCALA']
    
    for t in textos:
        c = t.get('conteudo', '')
        if not c: continue
        c_upper = c.upper()
        if any(r in c_upper for r in ruidos): continue
        raw_pos = t.get('posicao', [0, 0])
        pos_xy = tuple(round(p, 1) for p in raw_pos[:2])
        conteudo_limpo = _clean_text_servicos(c)
        match_metragem = re.search(r'(\d+[.,]?\d*)\s*(m|metros)\b', conteudo_limpo, re.IGNORECASE)
        metragem_val = float(match_metragem.group(1).replace(',','.')) if match_metragem else None
        
        is_demolicao = any(k in c_upper for k in ['DEMOL', 'DEMOLIÇÃO', 'DEMOLIDA', 'DEMOLIDO'])
        is_remocao = any(k in c_upper for k in ['REMOV', 'REMOÇÃO', 'RETIRAR', 'REMOVIDO', 'REMOVIDA'])
        
        if is_demolicao:
            if c_upper.strip() not in ['A DEMOLIR', 'DEMOLIÇÃO', 'DEMOLIR', 'A SER DEMOLIDA', 'A SER DEMOLIDO']:
                sujeito = re.sub(r'\bA\s+(SER\s+)?DEMOL\w+\b(\s*/\s*REATERRAR)?', '', conteudo_limpo, flags=re.IGNORECASE)
                sujeito = re.sub(r'\(\s*\)', '', sujeito)
                sujeito = re.sub(r'-\s*$', '', sujeito).strip()
                if len(sujeito) > 1:
                    nome_padronizado = f"{sujeito.upper()} A DEMOLIR"
                    chave = (nome_padronizado, pos_xy)
                    if chave not in seen_dem:
                        seen_dem.add(chave)
                        counter_dem[nome_padronizado] = counter_dem.get(nome_padronizado, 0) + 1
                        idx = counter_dem[nome_padronizado]
                        nome_exibicao = nome_padronizado if idx == 1 else f"{nome_padronizado} ({idx})"
                        is_continuo = any(k in nome_padronizado for k in ['DUTO', 'FIAÇÃO', 'CABO', 'ALVENARIA', 'PLATIBANDA', 'PISO'])
                        is_pilar_estrutural = any(k in nome_padronizado for k in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'PILAR', 'POSTE', 'FUNDAÇÃO'])
                        if is_continuo:
                            valor = metragem_val if metragem_val is not None else ""
                        elif is_pilar_estrutural:
                            valor = ""
                        else:
                            valor = 1
                        demolicoes_final.append((nome_exibicao, valor))
                        
        if is_remocao and not is_demolicao:
            if c_upper.strip() not in ['A REMOVER', 'REMOÇÃO', 'REMOVER', 'A RETIRAR', 'RETIRAR']:
                sujeito = re.sub(r'\b(A\s+(SER\s+)?(REMOV\w+|RETIRAR\w*)|REMOÇÃO\s+DE|RETIRADA\s+DE)\b', '', conteudo_limpo, flags=re.IGNORECASE)
                sujeito = re.sub(r'\(\s*\)', '', sujeito)
                sujeito = re.sub(r'-\s*$', '', sujeito).strip()
                if len(sujeito) > 1:
                    nome_padronizado = f"{sujeito.upper()} A REMOVER"
                    chave = (nome_padronizado, pos_xy)
                    if chave not in seen_rem:
                        seen_rem.add(chave)
                        counter_rem[nome_padronizado] = counter_rem.get(nome_padronizado, 0) + 1
                        idx = counter_rem[nome_padronizado]
                        nome_exibicao = nome_padronizado if idx == 1 else f"{nome_padronizado} ({idx})"
                        is_continuo = any(k in nome_padronizado for k in ['DUTO', 'FIAÇÃO', 'FIACAO', 'CABO'])
                        if is_continuo:
                            valor = metragem_val if metragem_val is not None else ""
                        else:
                            valor = 1
                        remocoes_final.append((nome_exibicao, valor))
                        
    return remocoes_final, demolicoes_final

def escrever(ws, linha, coluna, valor):
    if valor is None or valor == "": return 
    celula = ws.cell(row=linha, column=coluna)
    if isinstance(celula, MergedCell):
        for rng in ws.merged_cells.ranges:
            if celula.coordinate in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = valor
                return
    else:
        celula.value = valor

def _escrever_linha_padrao_solo(ws, linha, nome, inclinacao, volume):
    escrever(ws, linha, SOLO_COL_AMB, nome)
    escrever(ws, linha, SOLO_COL_I,   inclinacao if nome else None)
    escrever(ws, linha, SOLO_COL_V,   volume if nome else None)

def _comp_fuzzy(entidades, palavra_chave):
    total = 0.0
    chave = palavra_chave.upper()
    for e in entidades:
        layer = e.get('layer', '').upper()
        if chave in layer or chave.replace('Ê', 'E').replace('Ç', 'C') in layer:
            d = e.get('dados', {})
            if e.get('tipo') == 'LINE':
                total += d.get('comprimento', 0.0)
            elif e.get('tipo') in ('LWPOLYLINE', 'POLYLINE'):
                pts = d.get('pontos', [])
                for i in range(len(pts) - 1):
                    total += math.sqrt((pts[i+1][0] - pts[i][0])**2 + (pts[i+1][1] - pts[i][1])**2)
    return round(total, 2)

def _comp_entidade(e):
    d = e.get('dados', {})
    tipo = e.get('tipo', '')
    if tipo == 'LINE':
        return d.get('comprimento', 0.0)
    if tipo in ('LWPOLYLINE', 'POLYLINE'):
        pts = d.get('pontos', [])
        total = 0.0
        for i in range(len(pts) - 1):
            total += math.sqrt((pts[i+1][0] - pts[i][0])**2 + (pts[i+1][1] - pts[i][1])**2)
        return round(total, 2)
    return 0.0

def _match_ambiente_fuzzy(descricao, nomes_amb):
    desc = re.sub(r'\s+', '', descricao.upper())
    melhor = None
    melhor_score = 0
    for nome in nomes_amb:
        nome_limpo = re.sub(r'\s+', '', nome.upper())
        if nome_limpo in desc:
            score = len(nome_limpo)
            if score > melhor_score:
                melhor = nome
                melhor_score = score
        elif desc in nome_limpo:
            score = len(desc)
            if score > melhor_score:
                melhor = nome
                melhor_score = score
        else:
            palavras = [p for p in nome.upper().split() if len(p) > 2]
            if palavras and all(p in descricao.upper() for p in palavras):
                score = sum(len(p) for p in palavras)
                if score > melhor_score:
                    melhor = nome
                    melhor_score = score
    return melhor

def _extrair_dxf_por_ambiente(entidades, textos, blocos, ambientes):
    if not ambientes:
        return {}

    estruturas_mapeadas = {a['nome']: {'pilares': []} for a in ambientes if a.get('nome')}
    
    for e in entidades:
        layer = e.get('layer', '').upper()
        if 'PILAR' in layer:
            d = e.get('dados', {})
            tipo = e.get('tipo', '')
            pos = None
            comp = _comp_entidade(e)
            is_poly = False
            c_med = 0.0
            l_med = 0.0
            
            if tipo == 'LINE':
                pos = d.get('inicio', [0, 0])
                c_med = comp
            elif tipo in ('LWPOLYLINE', 'POLYLINE'):
                pts = d.get('pontos', [])
                if pts:
                    pos = pts[0]
                    xs = [p[0] for p in pts]
                    ys = [p[1] for p in pts]
                    c_med = max(xs) - min(xs)
                    l_med = max(ys) - min(ys)
                    is_poly = True
            
            if pos and comp > 0:
                menor_dist = float('inf')
                amb_alvo = None
                for amb in ambientes:
                    nome = amb.get('nome')
                    if not nome or 'posicao' not in amb: continue
                    pos_amb = amb['posicao']
                    dist = math.sqrt((pos[0] - pos_amb[0])**2 + (pos[1] - pos_amb[1])**2)
                    if dist < menor_dist:
                        menor_dist = dist
                        amb_alvo = nome
                        
                if amb_alvo:
                    estruturas_mapeadas[amb_alvo]['pilares'].append({
                        'comp': comp,
                        'pos': pos,
                        'is_poly': is_poly,
                        'c': c_med,
                        'l': l_med
                    })

    padroes_ele = {
        'tomadas': r'\bTOM\.?\s*(?:AR|BAIXA|M[ÉE]DIA|ALTA)?\b|\bTOMADAS?\b|\bTUGS?\b|\bTUES?\b|\bTOMADA\s+DE\s+FOR[ÇC]A\b|\bTF\b',
        'interruptores': r'\bINTERRUPTORES?\b|\bINT\.?\b|\bCHAVES?\b|\bCHAVE\s+(?:SIMPLES|PARALELA|INTERMEDI[ÁA]RIA)\b',
        'luminarias': r'\bLUMIN[ÁA]RIAS?\b|\bLUM\.?\b|\bILUM(?:INA[ÇC][ÃA]O)?\.?\b|\bL[ÂA]MPADAS?\b|\bLAMP\.?\b|\bPLAFON\b|\bSPOT\b',
        'conduletes': r'\bCONDULET[ES]*\b|\bCX\.?\s*MOLDADA\b|\bCAIXA\s+MOLDADA\b',
    }

    ele_por_amb = {a['nome']: {'tomadas': 0, 'interruptores': 0, 'luminarias': 0, 'quadros': 0, 'conduletes': 0} for a in ambientes if a.get('nome')}
    nomes_amb = [a['nome'].upper() for a in ambientes if a.get('nome')]

    for t in textos:
        conteudo = _clean_text_ambientes(t.get('conteudo', '')).upper()
        if not conteudo:
            continue
        tem_ele = False
        for cat, padrao in padroes_ele.items():
            if re.search(padrao, conteudo):
                tem_ele = True
                break
        if not tem_ele:
            continue
        amb_match = _match_ambiente_fuzzy(conteudo, nomes_amb)
        if not amb_match:
            continue
        for cat, padrao in padroes_ele.items():
            if re.search(padrao, conteudo):
                if amb_match in ele_por_amb:
                    ele_por_amb[amb_match][cat] += 1
                break

    for b in blocos:
        attrs = b.get('atributos', {})
        if not attrs:
            continue
        placa = attrs.get('PLACA', '')
        amb_match = _match_ambiente_fuzzy(placa, nomes_amb)
        if amb_match and amb_match in ele_por_amb:
            ele_por_amb[amb_match]['quadros'] += 1

    duto_total = 0.0
    cabo_total = 0.0
    for e in entidades:
        layer = e.get('layer', '').upper()
        comp = _comp_entidade(e)
        if any(p in layer for p in ['DUTO', 'ELETRODUTO', 'ELETROCALHA', 'CONDUITE', 'CONDUTE', 'TUBULACAO', 'TUBULAÇÃO']):
            duto_total += comp
        elif any(p in layer for p in ['CABO', 'FIAÇÃO', 'FIACAO', 'REDE LÓGICA', 'REDE LOGICA', 'TELEFONIA', 'CFTV']):
            cabo_total += comp

    total_area = sum(a.get('area', 0) for a in ambientes if a.get('area', 0) > 0)
    if total_area == 0:
        total_area = 1

    por_ambiente = {}
    for amb in ambientes:
        nome = amb.get('nome', '')
        if not nome:
            continue
        area = amb.get('area', 0)
        frac = area / total_area if total_area > 0 else 1.0 / len(ambientes)
        h_amb = amb.get('h', PD_PADRAO)

        est_amb = estruturas_mapeadas.get(nome, {'pilares': []})
        
        pilares_info = []
        linhas_p = []
        
        for p in est_amb['pilares']:
            if p['is_poly'] and p['c'] > 0 and p['l'] > 0:
                pilares_info.append({'c': round(p['c'], 2), 'l': round(p['l'], 2), 'h': h_amb})
            else:
                linhas_p.append(p)
                
        visitados = set()
        for i, l1 in enumerate(linhas_p):
            if i in visitados: continue
            grupo = [l1['comp']]
            visitados.add(i)
            for j, l2 in enumerate(linhas_p):
                if j in visitados: continue
                dist = math.sqrt((l1['pos'][0] - l2['pos'][0])**2 + (l1['pos'][1] - l2['pos'][1])**2)
                if dist < 1.0:
                    grupo.append(l2['comp'])
                    visitados.add(j)
            
            c = round(max(grupo), 2)
            l = round(min(grupo), 2) if len(grupo) >= 2 else c
            pilares_info.append({'c': c, 'l': l, 'h': h_amb})

        pilares_agrupados = Counter((p['c'], p['l'], p['h']) for p in pilares_info)
        
        c_pilar = l_pilar = h_pilar = None
        
        if pilares_agrupados:
            (c, l, h), _ = pilares_agrupados.most_common(1)[0]
            c_pilar = c
            l_pilar = l
            h_pilar = h

        estr = {
            'pilares_c': c_pilar,
            'pilares_l': l_pilar,
            'pilares_h': h_pilar
        }

        ele = ele_por_amb.get(nome, {'tomadas': 0, 'interruptores': 0, 'luminarias': 0, 'quadros': 0, 'conduletes': 0})
        ele['dutos_m'] = round(duto_total * frac, 2)
        ele['cabos_m'] = round(cabo_total * frac, 2)

        por_ambiente[nome] = {'estrutura': estr, 'eletrica': ele}

    return por_ambiente

def _extrair_ambientes_super(textos):
    ambientes = []
    seen = set()
    vaos_detectados = []
    keywords = ['SALA', 'CIRCULAÇÃO', 'CENTRO', 'BANHEIRO', 'QUARTO', 'COZINHA', 'DEPÓSITO', 'ALOJAMENTO', 'AREA', 'PÁTIO', 'WC', 'COPA', 'AUDITÓRIO', 'TELHADO']
    for t in textos:
        c = t.get('conteudo', '').upper()
        if re.match(r'^[PJ][A-Z]?\d+$', c):
            vaos_detectados.append(t)
    for t in textos:
        c = t.get('conteudo', '')
        if not c: continue
        area_match = re.search(r'(\d+[.,]?\d*)\s*m[²2]', c, re.IGNORECASE)
        parts = c.split('\\P')
        area_idx = next((idx for idx, p in enumerate(parts) if re.search(r"m[²2]", p, re.IGNORECASE)), -1)
        nome_raw = ""
        if area_idx > 0:
            for p in parts[:area_idx]:
                p_c = _clean_text_ambientes(p)
                nome_raw += (" " + p_c) if nome_raw else p_c
        else:
            nome_raw = _clean_text_ambientes(parts[0])
        nome_final = nome_raw.upper().strip()
        if len(nome_final) < 2 or nome_final in seen: continue
        if area_match or any(k in nome_final for k in keywords):
            area_val = float(area_match.group(1).replace(',', '.')) if area_match else 0.0
            perim = re.search(r'P\s*=\s*(\d+[.,]?\d*)', c, re.IGNORECASE)
            pd = re.search(r'PD\s*=\s*(\d+[.,]?\d*)', c, re.IGNORECASE)
            v_perim = float(perim.group(1).replace(',', '.')) if perim else None
            v_pd = float(pd.group(1).replace(',', '.')) if pd else PD_PADRAO
            c_val = l_val = None
            if v_perim and area_val > 0:
                s = v_perim / 2
                disc = s**2 - 4 * area_val
                if disc >= 0:
                    c_val = round((s + math.sqrt(disc)) / 2, 2)
                    l_val = round((s - math.sqrt(disc)) / 2, 2)
            pos_amb = t.get('posicao', [0, 0])
            vaos_amb = []
            for v in vaos_detectados:
                pos_v = v.get('posicao', [0, 0])
                if math.sqrt((pos_amb[0] - pos_v[0])**2 + (pos_amb[1] - pos_v[1])**2) < RAIO_VAOS:
                    vaos_amb.append(v['conteudo'])
            ambientes.append({'nome': nome_final, 'area': area_val, 'perimetro': v_perim, 'c': c_val, 'l': l_val, 'h': v_pd, 'posicao': pos_amb, 'e': ESPESSURA_PADRAO, 'vaos': ", ".join(vaos_amb)})
            seen.add(nome_final)
    return ambientes

def _preencher_amb_base(ws, linha, amb: dict):
    escrever(ws, linha, COL_AMB_NOME, amb.get('nome'))
    escrever(ws, linha, COL_AMB_C, amb.get('comprimento') or amb.get('c'))
    escrever(ws, linha, COL_AMB_L, amb.get('largura') or amb.get('l'))
    escrever(ws, linha, COL_AMB_H, amb.get('altura') or amb.get('h', PD_PADRAO))
    escrever(ws, linha, COL_AMB_E, ESPESSURA_PADRAO)
    escrever(ws, linha, COL_AMB_AREA, amb.get('area'))
    vaos = amb.get('vaos') or ""
    if not vaos and (amb.get('janelas') or amb.get('portas')):
        partes = [f"{j.get('tipo','J')}({j.get('quantidade',1)}x)" for j in amb.get('janelas', [])] + [f"{p.get('tipo','P')}({p.get('quantidade',1)}x)" for p in amb.get('portas', [])]
        vaos = ", ".join(partes)
    escrever(ws, linha, COL_AMB_VAOS, vaos)

def _preencher_eletrica(ws, linha, amb: dict):
    escrever(ws, linha, 2, amb.get('nome'))
    escrever(ws, linha, COL_ELE_TOMADAS, amb.get('tomadas'))
    escrever(ws, linha, COL_ELE_INTERR, amb.get('interruptores'))
    escrever(ws, linha, COL_ELE_LUM, amb.get('iluminacao'))
    escrever(ws, linha, COL_ELE_CONDULE, amb.get('caixasInspecao'))
    tipo_acc = " | ".join(filter(None, [amb.get('tipoTomada'), amb.get('tipoInterruptor'), amb.get('tipoLuminaria')]))
    escrever(ws, linha, COL_ELE_TIPO_AC, tipo_acc)
    if amb.get('cabos'):
        escrever(ws, linha, COL_ELE_CABOS, ", ".join(f"{c.get('circuito','')} {c.get('secao','')}mm²" for c in amb.get('cabos')))
    if amb.get('cabeamentos'):
        escrever(ws, linha, COL_ELE_DUTOS, round(sum(float(cb.get('comprimento') or 0) for cb in amb.get('cabeamentos')), 2))
    if amb.get('disjuntores'):
        escrever(ws, linha, COL_ELE_EQUIP, ", ".join(f"{d.get('amperagem','?')}A×{d.get('quantidade',1)}" for d in amb.get('disjuntores')))
        escrever(ws, linha, COL_ELE_EQUIP_QTD, sum(int(d.get('quantidade') or 1) for d in amb.get('disjuntores')))

def _preencher_hidraulica(ws, linha, amb: dict):
    escrever(ws, linha, 2, amb.get('nome'))
    escrever(ws, linha, COL_HID_REGISTROS, amb.get('registros'))
    escrever(ws, linha, COL_HID_VALVULAS, amb.get('valvulas'))
    if amb.get('reservatorio'):
        escrever(ws, linha, COL_HID_RESERVAT, 1)
        escrever(ws, linha, COL_HID_RESERVAT_L, amb['reservatorio'].get('capacidade'))
    if amb.get('ramais'):
        escrever(ws, linha, COL_HID_AF_DUTOS, round(sum(float(r.get('comprimento') or 0) for r in amb.get('ramais')), 2))

def _preencher_rede_spda(ws, linha, amb: dict):
    escrever(ws, linha, 2, amb.get('nome'))
    escrever(ws, linha, COL_REDE_QUADROS, amb.get('quadrosRede'))
    escrever(ws, linha, COL_REDE_TOMADAS, amb.get('patchCords'))
    escrever(ws, linha, COL_REDE_CONDULE, amb.get('cameras'))
    escrever(ws, linha, COL_SPDA_CAPTACAO, amb.get('terminaisAereos'))
    escrever(ws, linha, COL_SPDA_CONDULE, amb.get('caixasInspecao'))
    escrever(ws, linha, COL_SPDA_ATERR, amb.get('hastesAterramento'))

def _preencher_incendio(ws, linha, amb: dict):
    escrever(ws, linha, 2, amb.get('nome'))
    ext = amb.get('extintores', [])
    if ext:
        escrever(ws, linha, COL_INC_RES_TIPO, ", ".join(f"{e.get('tipo','')} {e.get('peso','')}kg" for e in ext))
        escrever(ws, linha, COL_INC_RES_QTD, len(ext))
    hid = amb.get('hidrantes', [])
    if hid:
        escrever(ws, linha, COL_INC_HIDRANTES, len(hid))
        escrever(ws, linha, COL_INC_DUTOS, round(sum(float(h.get('comprimento') or 0) for h in hid), 2))
        diams = list({h.get('diametro') for h in hid if h.get('diametro')})
        if diams: escrever(ws, linha, COL_INC_DUTOS_TIPO, ", ".join(str(d) for d in diams))

def _preencher_cobertura(ws, linha, amb: dict):
    escrever(ws, linha, 2, amb.get('nome'))
    escrever(ws, linha, COL_COB_C, amb.get('comprimento') or amb.get('c'))
    escrever(ws, linha, COL_COB_L, amb.get('largura') or amb.get('l'))
    escrever(ws, linha, COL_COB_H, amb.get('altura') or amb.get('h'))
    escrever(ws, linha, COL_COB_EST_TIPO, amb.get('tipoEstrutura'))
    escrever(ws, linha, COL_COB_TEL_TIPO, amb.get('tipoTelhamento'))
    escrever(ws, linha, COL_COB_TEL_E, amb.get('espessura'))
    escrever(ws, linha, COL_COB_EST_E, amb.get('inclinacao'))
    if amb.get('pecas'):
        desc = ", ".join(f"{p.get('descricao','')} {p.get('secao','')}".strip() for p in amb['pecas'])
        escrever(ws, linha, COL_COB_EST_TIPO, f"{amb.get('tipoEstrutura','')} | {desc}".strip(' |'))

def mesclar_form_com_dxf(dados_form: dict | list, dados_dxf: dict) -> dict:
    entidades = dados_dxf.get('entidades', [])
    if not entidades:
        for val in dados_dxf.values():
            if isinstance(val, dict) and 'entidades' in val:
                entidades = val['entidades']
                break

    textos = dados_dxf.get('textos', []) or [
        e for e in entidades if e.get('tipo') in ('MTEXT', 'TEXT')
    ]

    if 'ambientes' not in dados_form:
        ambiente = {k: v for k, v in dados_form.items() if k != 'projeto_id'}
        dados_form = {'ambientes': [ambiente]}

    ambientes_dxf = _extrair_ambientes_super(textos)
    ambientes_dxf_validos = [a for a in ambientes_dxf if a.get('area', 0) > 0 or a.get('comprimento') or a.get('c')]

    nomes_form = {a.get('nome', '').upper().strip() for a in dados_form.get('ambientes', [])}
    ambientes_novos = [a for a in ambientes_dxf_validos if a.get('nome', '').upper().strip() not in nomes_form]

    resultado = dict(dados_form)
    resultado['ambientes'] = list(dados_form.get('ambientes', [])) + ambientes_novos
    resultado['entidades'] = entidades
    resultado['textos'] = textos
    resultado['blocos'] = dados_dxf.get('blocos', [])
    return resultado

def extrair_memorial_calculo(dados_json: dict, ambiente_obj=None, template_path: str = None) -> bytes:
    is_form = 'ambientes' in dados_json

    if not is_form:
        if 'entidades' not in dados_json:
            for val in dados_json.values():
                if isinstance(val, dict) and 'entidades' in val:
                    dados_json = val; break
                if isinstance(val, str):
                    try:
                        parsed = json.loads(val)
                        if 'entidades' in parsed: dados_json = parsed; break
                    except Exception: pass

    entidades = dados_json.get('entidades', [])
    textos = dados_json.get('textos', []) or [e for e in entidades if e.get('tipo') in ('MTEXT', 'TEXT')]
    blocos = dados_json.get('blocos', [])

    ambientes = dados_json['ambientes'] if is_form else _extrair_ambientes_super(textos)
    dxf_por_amb = _extrair_dxf_por_ambiente(entidades, textos, blocos, ambientes) if entidades else {}

    eletrica  = _comp_fuzzy(entidades, 'FIAÇÃO') + _comp_fuzzy(entidades, 'ELETRODUTO')
    hidro_af  = _comp_fuzzy(entidades, 'ÁGUA FRIA') + _comp_fuzzy(entidades, 'HIDRAULICA')
    incendio  = _comp_fuzzy(entidades, 'INCÊNDIO') + _comp_fuzzy(entidades, 'HIDRANTE')
    itens_remocao, itens_demolicao = _extrair_servicos_agrupados(textos)

    tpl = Path(template_path) if template_path else CAMINHO_TEMPLATE
    wb  = openpyxl.load_workbook(str(tpl))
    
    ws_lev = wb['Levantamento Campo']
    ws_serv = wb['Serviços Preliminares']
    ws_solo = wb['Movimento de Solo'] 

    for nome_aba in wb.sheetnames:
        if nome_aba == 'Levantamento Campo': continue
        aba = wb[nome_aba]
        for row in aba.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("='Levantamento Campo'"):
                    cell.value = None

    for i, amb in enumerate(ambientes[:MAX_AMBIENTES]):
        r_amb, r_ele, r_hid, r_red, r_inc, r_cob, r_est = (
            LINHA_INI_AMBIENTES + i, LINHA_INI_ELETRICA + i, LINHA_INI_HIDRAULICA + i,
            LINHA_INI_REDE + i, LINHA_INI_INCENDIO + i, LINHA_INI_COBERTURA + i, LINHA_INI_ESTRUTURAS + i
        )

        _preencher_amb_base(ws_lev, r_amb, amb)

        nome_amb = (amb.get('nome') or '').upper().strip()
        dxf_data = dxf_por_amb.get(nome_amb, {})
        dxf_est = dxf_data.get('estrutura', {}) 
        dxf_ele = dxf_data.get('eletrica', {})

        if is_form:
            _preencher_eletrica(ws_lev, r_ele, amb)
            _preencher_hidraulica(ws_lev, r_hid, amb)
            _preencher_rede_spda(ws_lev, r_red, amb)
            _preencher_incendio(ws_lev, r_inc, amb)
            _preencher_cobertura(ws_lev, r_cob, amb)
        else:
            escrever(ws_lev, r_ele, 2, amb.get('nome'))
            escrever(ws_lev, r_hid, 2, amb.get('nome'))
            escrever(ws_lev, r_red, 2, amb.get('nome'))
            escrever(ws_lev, r_inc, 2, amb.get('nome'))

        tem_pilares = dxf_est.get('pilares_c') or dxf_est.get('pilares_l')
        if tem_pilares:
            escrever(ws_lev, r_est, 2, amb.get('nome'))
            escrever(ws_lev, r_est, COL_EST_PILAR_C, dxf_est.get('pilares_c'))
            escrever(ws_lev, r_est, COL_EST_PILAR_L, dxf_est.get('pilares_l'))
            escrever(ws_lev, r_est, COL_EST_PILAR_H, dxf_est.get('pilares_h'))
        else:
            escrever(ws_lev, r_est, 2, "")

        if not amb.get('tomadas'):
            escrever(ws_lev, r_ele, COL_ELE_QUADROS, dxf_ele.get('quadros') or None)
            escrever(ws_lev, r_ele, COL_ELE_CONDULE, dxf_ele.get('conduletes') or None)
            escrever(ws_lev, r_ele, COL_ELE_TOMADAS, dxf_ele.get('tomadas') or None)
            escrever(ws_lev, r_ele, COL_ELE_INTERR, dxf_ele.get('interruptores') or None)
            escrever(ws_lev, r_ele, COL_ELE_LUM, dxf_ele.get('luminarias') or None)
            escrever(ws_lev, r_ele, COL_ELE_DUTOS, dxf_ele.get('dutos_m') or None)
            escrever(ws_lev, r_ele, COL_ELE_CABOS, dxf_ele.get('cabos_m') or None)

        area_s, inc_s, prof_s = float(amb.get('area') or 0.0), float(amb.get('inclinacaoTerreno') or 0.0), float(amb.get('profundidadeEscavacao') or 0.0)
        vols = amb.get('volumes') or {}
        v_terra, v_ater, v_enro, v_cont, v_talu, v_niv, v_comp = float(vols.get('terraplanagem') or 0.0), float(vols.get('aterro') or 0.0), float(vols.get('enrocamento') or 0.0), float(vols.get('contencao') or 0.0), float(vols.get('taludamento') or 0.0), float(vols.get('nivelamento') or 0.0), float(vols.get('compactacao') or 0.0)
        v_esc = float(vols.get('escavacao') or 0.0) or round(area_s * prof_s, 2)

        l_solo_terra = SOLO_LINHA_TERRAPLANAGEM + i
        if v_terra > 0 or v_esc > 0 or prof_s > 0:
            escrever(ws_solo, l_solo_terra, TERRA_COL_AMB, amb.get('nome'))
            escrever(ws_solo, l_solo_terra, TERRA_COL_TIPO, 'Corte')
            escrever(ws_solo, l_solo_terra, TERRA_COL_H, prof_s if prof_s > 0 else None)
            escrever(ws_solo, l_solo_terra, TERRA_COL_V, v_esc if v_esc > 0 else v_terra)
        else: escrever(ws_solo, l_solo_terra, TERRA_COL_AMB, None)

        _escrever_linha_padrao_solo(ws_solo, SOLO_LINHA_ATERRO + i, amb.get('nome') if v_ater > 0 else None, inc_s, v_ater)
        _escrever_linha_padrao_solo(ws_solo, SOLO_LINHA_ENROCAMENTO + i, amb.get('nome') if v_enro > 0 else None, None, v_enro)
        _escrever_linha_padrao_solo(ws_solo, SOLO_LINHA_CONTENCAO + i, amb.get('nome') if v_cont > 0 else None, None, v_cont)
        _escrever_linha_padrao_solo(ws_solo, SOLO_LINHA_TALUDAMENTO + i, amb.get('nome') if v_talu > 0 else None, None, v_talu)
        v_total_n = round(v_niv + v_comp, 3)
        _escrever_linha_padrao_solo(ws_solo, SOLO_LINHA_NIVELAMENTO + i, amb.get('nome') if v_total_n > 0 else None, None, v_total_n)

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    log_line = f"[AUTO] {ts} | {len(ambientes)} amb | Fiação={eletrica}m | Incêndio={incendio}m"
    escrever(ws_lev, LINHA_LOG, 2, log_line)

    l_ini_rem, l_ini_dem = None, None
    for r in range(1, 200):
        val_l = str(ws_serv.cell(row=r, column=2).value or "").lower()
        if "remoções" in val_l: l_ini_rem = r + 5
        elif "demolições" in val_l: l_ini_dem = r + 6

    if l_ini_rem:
        for i, (nome, valor) in enumerate(itens_remocao[:MAX_ITENS]):
            linha = l_ini_rem + i
            escrever(ws_serv, linha, 2, nome)
            col_id = get_coluna_remocao(nome)
            escrever(ws_serv, linha, 29 if col_id == "GENERICO" else col_id, valor)

    if l_ini_dem:
        for i, (nome, valor) in enumerate(itens_demolicao[:MAX_ITENS]):
            linha = l_ini_dem + i
            escrever(ws_serv, linha, 2, nome)
            col_id = get_coluna_demolicao(nome)
            if col_id: escrever(ws_serv, linha, col_id, valor)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()