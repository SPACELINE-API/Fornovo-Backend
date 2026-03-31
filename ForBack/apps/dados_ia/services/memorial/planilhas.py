import os
import re
import math
from pathlib import Path
from datetime import datetime
import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
PASTA_EXPORTACAO = BASE_DIR / "planilhas_geradas"
CAMINHO_TEMPLATE = Path(__file__).resolve().parent / "templates_excel" / "Memorial de Cálculo - Modelo.xlsx"

ELE_LAYERS = {
    'fiacao': ['ELE-FIAÇÃO', 'ELE-FIAÇÃO1', 'ELE-FIACAO', 'ELE-FIACAO1'],
    'circuito': ['ELE-CIRCUITO'],
    'eletrocalha': ['ELE-ELETROCALHA'],
    'malha_terra': ['ELE-MALHA TERRA'],
    'malha_spda': ['ELE-MALHA SPDA'],
    'barra_chata': ['ELE-BARRA CHATA ALUMÍNIO'],
    'cordoalha': ['ELE-CORDOALHA'],
    'terminal_aereo': ['ELE-TERMINAL AÉREO'],
    'textos': ['ELE-TEXTOS', 'ELE-TEXTO5'],
    'eletrica': ['Elétrica'],
    'logica': ['Lógica - Dados Telefonia CFTV'],
}


def _comprimento_entidades(entidades, layers):
    total = 0.0
    for e in entidades:
        if e.get('layer') not in layers:
            continue
        d = e.get('dados', {})
        tipo = e.get('tipo')
        if tipo == 'LINE':
            total += d.get('comprimento', 0.0)
        elif tipo == 'LWPOLYLINE':
            pts = d.get('pontos', [])
            for i in range(len(pts) - 1):
                dx = pts[i+1][0] - pts[i][0]
                dy = pts[i+1][1] - pts[i][1]
                total += math.sqrt(dx**2 + dy**2)
        elif tipo == 'POLYLINE':
            pts = d.get('pontos', [])
            for i in range(len(pts) - 1):
                dx = pts[i+1][0] - pts[i][0]
                dy = pts[i+1][1] - pts[i][1]
                dz = pts[i+1][2] - pts[i][2]
                total += math.sqrt(dx**2 + dy**2 + dz**2)
    return round(total, 2)


def _contar_entidades(entidades, layers):
    return sum(1 for e in entidades if e.get('layer') in layers)


def _extrair_ambientes(textos):
    ambientes = []
    seen = set()
    for t in textos:
        if t.get('layer') != 'ARQ - Textos':
            continue
        c = t['conteudo']
        area_m = re.search(r'(\d+[.,]\d+)m2', c.replace('²', '2'))
        if not area_m:
            continue
        parts = c.split('\\P')
        nome_raw = parts[0] if parts else ''
        nome = re.sub(r'\\[a-zA-Z]+[^;]*;', '', nome_raw).strip().strip('\\{').strip()
        perim_m = re.search(r'P=(\d+[.,]\d+)m', c)
        if nome and nome not in seen and len(nome) > 2 and not any(x in nome.upper() for x in ['TELHADO', 'PÁTIO', 'CALÇADA']):
            area = float(area_m.group(1).replace(',', '.'))
            perim = float(perim_m.group(1).replace(',', '.')) if perim_m else None
            ambientes.append({
                'nome': nome,
                'area': area,
                'perimetro': perim,
                'tug': math.ceil(area / 5),
                'iluminacao': math.ceil(area / 10),
                'interruptores': math.ceil(area / 10),
            })
            seen.add(nome)
    return ambientes


def _extrair_circuitos(textos):
    circuitos = set()
    for t in textos:
        if t.get('layer') == 'ELE-CIRCUITO':
            matches = re.findall(r'C[SP]?-?\d+', t['conteudo'])
            circuitos.update(matches)
    return sorted(circuitos)


def _extrair_ar_condicionado(textos):
    btu_counts = {}
    cortinas = 0
    for t in textos:
        if t.get('layer') != 'Elétrica':
            continue
        c = t['conteudo'].upper()
        matches = re.findall(r'(\d+[\.,]?\d*)\s*BTU', c)
        for v in matches:
            key = v.replace('.', '').replace(',', '')
            btu_counts[key] = btu_counts.get(key, 0) + 1
        if 'CORTINA' in c:
            cortinas += 1
    return btu_counts, cortinas


def _extrair_aterramento(textos, entidades):
    haste_textos = [t for t in textos if 'HASTE' in t.get('conteudo', '').upper()
                    and t.get('layer', '').startswith('ELE')]
    haste_unicas = set()
    for t in haste_textos:
        c = t['conteudo'].upper()
        if 'HASTE COBREADA' in c or 'HASTE DE TERRA' in c or 'HASTE DE ATERRAMENTO' in c:
            haste_unicas.add(t['conteudo'].strip())
    qtd_hastes = len(haste_unicas)

    cx_textos = [t for t in textos if 'INSPE' in t.get('conteudo', '').upper()
                 and ('CAIXA' in t.get('conteudo', '').upper() or 'CX' in t.get('conteudo', '').upper())
                 and t.get('layer', '').startswith('ELE')]
    cx_unicas = set(t['conteudo'].strip() for t in cx_textos if 'CAIXA DE INSPEÇÃO' in t['conteudo'].upper())
    qtd_cx = len(cx_unicas)

    condutor_cobre = None
    condutor_spda_cu = None
    condutor_spda_bc = None
    for t in textos:
        c = t.get('conteudo', '')
        if 'CORDOALHA' in c.upper() and 'S=' in c.upper():
            m = re.search(r'S=(\d+[.,]\d+)', c)
            if m:
                condutor_cobre = f"Cobre nú S={m.group(1)}mm²"
        if 'CONDUTOR PARA SPDA, EM COBRE' in c.upper():
            m = re.search(r'S=(\d+[.,]\d+)', c)
            if m:
                condutor_spda_cu = f"Cobre nú S={m.group(1)}mm²"
        if 'BARRA CHATA DE ALUMÍNIO' in c.upper() and '#' in c:
            m = re.search(r'#(\d+)mm', c)
            esp = re.search(r'ESPESSURA DE (\d+)mm', c.upper())
            if m:
                condutor_spda_bc = f"Barra chata Al #{m.group(1)}mm² {('esp. '+esp.group(1)+'mm') if esp else ''}"

    comp_malha_terra = _comprimento_entidades(entidades, ELE_LAYERS['malha_terra'])
    comp_malha_spda = _comprimento_entidades(entidades, ELE_LAYERS['malha_spda'])
    comp_barra_chata = _comprimento_entidades(entidades, ELE_LAYERS['barra_chata'])
    comp_cordoalha = _comprimento_entidades(entidades, ELE_LAYERS['cordoalha'])
    qtd_terminais = _contar_entidades(entidades, ELE_LAYERS['terminal_aereo'])

    return {
        'qtd_hastes': qtd_hastes,
        'qtd_cx_inspecao': qtd_cx,
        'qtd_terminais_aereos': qtd_terminais,
        'condutor_cobre_nu': condutor_cobre,
        'condutor_spda_cobre': condutor_spda_cu,
        'condutor_spda_barra_chata': condutor_spda_bc,
        'comp_malha_terra_m': comp_malha_terra,
        'comp_malha_spda_m': comp_malha_spda,
        'comp_barra_chata_m': comp_barra_chata,
        'comp_cordoalha_m': comp_cordoalha,
    }


def _extrair_qdg(textos):
    qdg_textos = [t for t in textos if 'QDG' in t.get('conteudo', '').upper()]
    quadros = set()
    for t in qdg_textos:
        c = t['conteudo']
        m = re.findall(r'QDG[\s/\\P]*\d*', c.upper())
        for q in m:
            quadros.add(q.strip())
    return list(quadros)


def _extrair_logica(textos):
    log_textos = [t for t in textos if t.get('layer') == 'Lógica - Dados Telefonia CFTV']
    eletrocalha = None
    eletrodutos = []
    saidas_rj45 = 0
    cameras = 0
    for t in log_textos:
        c = t['conteudo']
        if 'ELETROCALHA' in c.upper():
            eletrocalha = c.strip()
        if 'ELETRODUTO' in c.upper():
            eletrodutos.append(c.strip())
        if 'RJ' in c.upper() or 'RJ45' in c.upper():
            saidas_rj45 += 1
        if 'CÂMERA' in c.upper() or 'CAMERA' in c.upper() or 'CFTV' in c.upper():
            cameras += 1
    return {
        'eletrocalha_logica': eletrocalha,
        'eletrodutos_logica': list(set(eletrodutos)),
        'saidas_rj45': saidas_rj45,
        'cameras_cftv': cameras,
    }


def extrair_dados_eletricos_para_xlsx(dados_json: dict) -> str:
    textos = dados_json.get('textos', [])
    entidades = dados_json.get('entidades', [])

    ambientes = _extrair_ambientes(textos)
    circuitos = _extrair_circuitos(textos)
    btu_counts, cortinas = _extrair_ar_condicionado(textos)
    aterramento = _extrair_aterramento(textos, entidades)
    quadros = _extrair_qdg(textos)
    logica = _extrair_logica(textos)

    comp_fiacao = _comprimento_entidades(entidades, ELE_LAYERS['fiacao'])
    comp_circuito = _comprimento_entidades(entidades, ELE_LAYERS['circuito'])

    template_path = CAMINHO_TEMPLATE
    if not template_path.exists():
        raise FileNotFoundError(f"Template não encontrado: {template_path}")

    wb = openpyxl.load_workbook(str(template_path))
    ws = wb['Inst Elétricas']

    # ── CIRCUITOS INTERNOS (rows 58–74, seção começa na row 52) ──────────────
    # Estrutura da planilha (row 55-57 = headers):
    # Col B=Local, E=Circuito, F=Tipo QDG, G=QtdDisjuntores
    # K=Cabo1.5, L=Cabo2.5, M=Cabo4, N=Cabo10, O=Cabo16
    # P=Eletrocalha L, Q=Eletrocalha h, R=Eletrocalha C
    # V=Eletroduto DN, W=Eletroduto C
    # AG=Tomadas Tipo, AH=Tomadas Qtd
    # AI=Interruptores Tipo, AJ=Interruptores Qtd
    # AK=Luminarias Tipo, AL=Luminarias Qtd
    # AM=Haste DN, AN=Haste C, AO=Haste Qtd
    # AQ=Cx Inspecao Qtd

    linha_inicio = 58
    for i, amb in enumerate(ambientes[:17]):
        row = linha_inicio + i
        ws.cell(row=row, column=2, value=amb['nome'])
        ws.cell(row=row, column=33, value='TUG')
        ws.cell(row=row, column=33, value=amb['tug'])
        ws.cell(row=row, column=35, value=amb['interruptores'])
        ws.cell(row=row, column=37, value=amb['iluminacao'])

    # Linha de totais do aterramento (usa primeira linha disponível)
    row_aterre = linha_inicio
    ws.cell(row=row_aterre, column=2, value='Aterramento / SPDA')
    ws.cell(row=row_aterre, column=38, value='5/8"')
    ws.cell(row=row_aterre, column=39, value=round(aterramento['comp_malha_terra_m'], 2))
    ws.cell(row=row_aterre, column=40, value=aterramento['qtd_hastes'] or None)
    ws.cell(row=row_aterre, column=41, value=aterramento['qtd_cx_inspecao'] or None)

    # ── CIRCUITOS SUBTERRÂNEOS (rows 32–48) ──────────────────────────────────
    # Col B=Local, E=Circuito, G=Cabo1.5 ... M=Cabo100
    # N=Eletroduto Metálico DN, O=Eletroduto Metálico C
    # P=Eletroduto Corrugado DN, Q=Eletroduto Corrugado C

    ws.cell(row=32, column=2, value='Fiação interna (ELE-FIAÇÃO)')
    ws.cell(row=32, column=5, value='Circuitos internos')
    ws.cell(row=32, column=7, value=round(comp_fiacao, 2))

    ws.cell(row=33, column=2, value='Circuitos elétricos (ELE-CIRCUITO)')
    ws.cell(row=33, column=5, value=f'{len(circuitos)} circuitos')
    ws.cell(row=33, column=7, value=round(comp_circuito, 2))

    if logica['eletrodutos_logica']:
        ws.cell(row=34, column=2, value='Lógica / CFTV')
        ws.cell(row=34, column=5, value='; '.join(logica['eletrodutos_logica'][:2]))

    # ── CIRCUITOS AÉREOS (rows 7–23) ─────────────────────────────────────────
    # Col B=Local, E=Circuito, G=Cabo1.5 ... M=Cabo100
    ws.cell(row=7, column=2, value='Padrão Entrada BT')
    ws.cell(row=7, column=5, value='Padrão de Entrada Baixa Tensão 1, 2, 3')

    ws.cell(row=8, column=2, value='SPDA — Barra Chata Al')
    ws.cell(row=8, column=5, value=aterramento['condutor_spda_barra_chata'] or '')
    ws.cell(row=8, column=13, value=round(aterramento['comp_barra_chata_m'], 2))

    ws.cell(row=9, column=2, value='SPDA — Cordoalha Cobre Nú')
    ws.cell(row=9, column=5, value=aterramento['condutor_spda_cobre'] or '')
    ws.cell(row=9, column=13, value=round(aterramento['comp_cordoalha_m'], 2))

    ws.cell(row=10, column=2, value='Malha Terra — Cobre Nú')
    ws.cell(row=10, column=5, value=aterramento['condutor_cobre_nu'] or '')
    ws.cell(row=10, column=13, value=round(aterramento['comp_malha_terra_m'], 2))

    ws.cell(row=11, column=2, value='Malha SPDA')
    ws.cell(row=11, column=13, value=round(aterramento['comp_malha_spda_m'], 2))

    # AR CONDICIONADO — info na área de circuitos aéreos
    row_ac = 13
    for btu, qtd in btu_counts.items():
        ws.cell(row=row_ac, column=2, value=f'Ar Cond. {btu} BTU')
        ws.cell(row=row_ac, column=5, value=f'{qtd} unidades')
        row_ac += 1
    if cortinas:
        ws.cell(row=row_ac, column=2, value='Cortina de ar')
        ws.cell(row=row_ac, column=5, value=f'{cortinas} unidades')

    PASTA_EXPORTACAO.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nome_arquivo = f'memorial_eletrico_{timestamp}.xlsx'
    caminho_xlsx = PASTA_EXPORTACAO / nome_arquivo
    wb.save(str(caminho_xlsx))

    return str(caminho_xlsx)


def extrair_dados_eletricos_para_csv(dados_json: dict) -> str:
    return extrair_dados_eletricos_para_xlsx(dados_json)