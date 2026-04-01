import io
from pathlib import Path
import openpyxl
from openpyxl.cell.cell import MergedCell

# Define o caminho para o template do Excel
CAMINHO_TEMPLATE = Path(__file__).resolve().parent / "templates_excel" / "Memorial de Cálculo - Modelo.xlsx"

MAX_AMBIENTES = 80

# --- MAPEAMENTO DAS LINHAS INICIAIS DE CADA TABELA ---
SOLO_LINHA_TERRAPLANAGEM = 17   
SOLO_LINHA_ATERRO        = 43  
SOLO_LINHA_ENROCAMENTO   = 69 
SOLO_LINHA_CONTENCAO     = 95  
SOLO_LINHA_TALUDAMENTO   = 121  
SOLO_LINHA_NIVELAMENTO   = 148  

# --- MAPEAMENTO DE COLUNAS (Tabela Terraplanagem/Corte) ---
TERRA_COL_AMB    = 2
TERRA_COL_TIPO   = 5
TERRA_COL_H      = 9
TERRA_COL_LASTRO = 10
TERRA_COL_V      = 12

# --- MAPEAMENTO DE COLUNAS (Restantes Tabelas) ---
SOLO_COL_AMB = 2
SOLO_COL_I   = 5
SOLO_COL_V   = 10


def _escrever(ws, linha, coluna, valor):
    """
    Função auxiliar que escreve no Excel lidando com células mescladas.
    Se o valor for None, limpa a célula (útil para limpar lixo do template).
    """
    if valor is None:
        # Quando passamos None, garantimos que a célula fica vazia
        valor_final = ""
    else:
        valor_final = valor

    celula = ws.cell(row=linha, column=coluna)
    if isinstance(celula, MergedCell):
        for rng in ws.merged_cells.ranges:
            if celula.coordinate in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = valor_final
                return
    else:
        celula.value = valor_final


def _escrever_linha_padrao(ws, linha, nome, inclinacao, volume):
    """
    Preenche as tabelas que seguem o padrão comum: 
    Nome do Ambiente | Inclinação | Volume
    Se 'nome' for None, passa None também para as outras para limpar a linha.
    """
    _escrever(ws, linha, SOLO_COL_AMB, nome)
    # Se o nome é None (vazio), as outras colunas também devem ser limpas
    _escrever(ws, linha, SOLO_COL_I,   inclinacao if nome else None)
    _escrever(ws, linha, SOLO_COL_V,   volume if nome else None)


def preencher_movimento_solo(wb, ambientes: list):
    """
    Lê a lista de ambientes (do JSON do Front-End) e distribui os volumes 
    e dados do terreno nas respectivas tabelas da aba Movimento de Solo.
    """
    if 'Movimento de Solo' not in wb.sheetnames:
        return

    ws = wb['Movimento de Solo']

    for i, amb in enumerate(ambientes[:MAX_AMBIENTES]):
        nome       = amb.get('nome', '')
        area       = float(amb.get('area') or 0.0)
        inclinacao = float(amb.get('inclinacaoTerreno')   or 0.0)
        prof_escav = float(amb.get('profundidadeEscavacao') or 0.0)
        
        volumes    = amb.get('volumes') or {}

        # 1. Extração dos volumes declarados no JSON
        v_terra  = float(volumes.get('terraplanagem') or 0.0)
        v_aterro = float(volumes.get('aterro')        or 0.0)
        v_enroc  = float(volumes.get('enrocamento')   or 0.0)
        v_cont   = float(volumes.get('contencao')     or 0.0)
        v_talud  = float(volumes.get('taludamento')   or 0.0)
        v_nivel  = float(volumes.get('nivelamento')   or 0.0)
        v_comp   = float(volumes.get('compactacao')   or 0.0)
        
        # 2. Fallback de Escavação (Calcula Área * Profundidade se o JSON não trouxer)
        v_escav_json = float(volumes.get('escavacao') or 0.0)
        if v_escav_json > 0:
            v_escav = v_escav_json
        else:
            v_escav = round(area * prof_escav, 2)

        # 3. Distribuição dos dados pelas Tabelas do Excel
        
        # --- Tabela 1: Terraplanagem / Escavação ---
        linha_terra = SOLO_LINHA_TERRAPLANAGEM + i
        if v_terra > 0 or v_escav > 0 or prof_escav > 0:
            _escrever(ws, linha_terra, TERRA_COL_AMB,  nome)
            _escrever(ws, linha_terra, TERRA_COL_TIPO, 'Corte')
            _escrever(ws, linha_terra, TERRA_COL_H,    prof_escav if prof_escav > 0 else None)
            _escrever(ws, linha_terra, TERRA_COL_V,    v_escav if v_escav > 0 else v_terra) 
        else:
            # Apaga TUDO na linha de Terraplanagem se não houver dados
            _escrever(ws, linha_terra, TERRA_COL_AMB,  None)
            _escrever(ws, linha_terra, TERRA_COL_TIPO, None)
            _escrever(ws, linha_terra, TERRA_COL_H,    None)
            _escrever(ws, linha_terra, TERRA_COL_V,    None)

        # --- Tabela 2: Aterro ---
        linha_aterro = SOLO_LINHA_ATERRO + i
        if v_aterro > 0:
            _escrever_linha_padrao(ws, linha_aterro, nome, inclinacao, v_aterro)
        else:
            _escrever_linha_padrao(ws, linha_aterro, None, None, None)

        # --- Tabela 3: Enrocamento ---
        linha_enroc = SOLO_LINHA_ENROCAMENTO + i
        if v_enroc > 0:
            _escrever_linha_padrao(ws, linha_enroc, nome, None, v_enroc)
        else:
            _escrever_linha_padrao(ws, linha_enroc, None, None, None)

        # --- Tabela 4: Contenção ---
        linha_cont = SOLO_LINHA_CONTENCAO + i
        if v_cont > 0:
            _escrever_linha_padrao(ws, linha_cont, nome, None, v_cont)
        else:
            _escrever_linha_padrao(ws, linha_cont, None, None, None)

        # --- Tabela 5: Taludamento ---
        linha_talud = SOLO_LINHA_TALUDAMENTO + i
        if v_talud > 0:
            _escrever_linha_padrao(ws, linha_talud, nome, None, v_talud)
        else:
            _escrever_linha_padrao(ws, linha_talud, None, None, None)

        # --- Tabela 6: Nivelamento / Compactação ---
        linha_nivel = SOLO_LINHA_NIVELAMENTO + i
        v_nivel_total = round(v_nivel + v_comp, 3)
        if v_nivel_total > 0:
            _escrever_linha_padrao(ws, linha_nivel, nome, None, v_nivel_total)
        else:
            _escrever_linha_padrao(ws, linha_nivel, None, None, None)


def extrair_movimento_solo(dados_json: dict, template_path: str = None) -> bytes:
    """
    Função principal a ser chamada pela View. 
    Abre o Excel, preenche a aba de Movimento de Solo e retorna os bytes.
    """
    ambientes = dados_json.get('ambientes', [])
    
    # Resolve o caminho do template (usa o path fornecido ou o padrão)
    tpl = Path(template_path) if template_path else CAMINHO_TEMPLATE
    
    # Carrega o workbook
    wb = openpyxl.load_workbook(str(tpl))
    
    # Executa a lógica de preenchimento
    preencher_movimento_solo(wb, ambientes)
    
    # Salva o resultado em memória e retorna como bytes
    output = io.BytesIO()
    wb.save(output)
    
    return output.getvalue()